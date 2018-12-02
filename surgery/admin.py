from django.contrib import admin
import nested_admin
from surgery.models import PatientBasic,CHA2DS2VACsEvaluation,HASBLEdEvaluation,UCG,TEE,InSurgery,Ablation,LAA,Antiarrhythmic,Anticoagulant,SurgeryMethod,AblationType,PreCheck,FollowTEE,FollowTTE,SurgeryFollow

class ModifiedTabInline(nested_admin.nested.NestedTabularInline):
    extra = 1
    max_num = 1
    can_delete = False

class cha2ds2vacsInline(ModifiedTabInline):
    model = CHA2DS2VACsEvaluation
    verbose_name = u'CHA2DS2-VACs评分'
    verbose_name_plural = verbose_name

class hasbledInline(ModifiedTabInline):
    model = HASBLEdEvaluation
    verbose_name_plural = u'HAS-BLED评分'
    verbose_name = verbose_name_plural

class UCGInline(ModifiedTabInline):
    model = UCG
    verbose_name_plural = u'UCG'

class TEEInline(ModifiedTabInline):
    model = TEE
    verbose_name_plural = u'TEE'


class AblationInline(ModifiedTabInline):
    model = Ablation
    verbose_name_plural = u'消融'

class LAAInline(ModifiedTabInline):
    model = LAA
    verbose_name_plural = u'左心耳封堵'

class PreCheckInline(ModifiedTabInline):
    model = PreCheck
    inlines = [UCGInline, TEEInline]
    verbose_name_plural = u'术前检查'
    verbose_name = verbose_name_plural


class InSurgeryInline(ModifiedTabInline):
    model = InSurgery
    inlines = [AblationInline, LAAInline]
    verbose_name_plural = u'术中信息'
    verbose_name = verbose_name_plural

class FollowTEEInline(ModifiedTabInline):
    model = FollowTEE

class FollowTTEInline(ModifiedTabInline):
    model = FollowTTE

class SurgeryFollowInline(nested_admin.nested.NestedTabularInline):
    model = SurgeryFollow
    inlines = [FollowTTEInline,FollowTEEInline]
    verbose_name = u'术后随访'
    verbose_name_plural = verbose_name
    extra = 4
    max_num = 4
    #can_delete = False

def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    #from openpyxl.cell import get_column_letter
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "MyModel"

    row_num = 0

    columns = [
        (u"序号", 15),
        (u"住院号", 15),
        (u"姓名", 70),
        (u"身份证号", 70),
        (u"性别", 15),
        (u"年龄", 15),
        (u"联系方式", 15),
        (u"房颤类型", 15),
        (u"其他疾病", 15),
        (u"生活质量评分", 15),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        # set column width
        #ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.id,
            obj.registerId,
            obj.name,
            obj.idstr,
            obj.gender,
            obj.age,
            obj.phone,
            obj.af,
            obj.otherDisease,
            obj.lifeQualityScore,
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"    

class PatientAdmin(nested_admin.nested.NestedModelAdmin):
    inlines = [cha2ds2vacsInline,hasbledInline,InSurgeryInline,PreCheckInline,SurgeryFollowInline]
    fieldsets = (
        (u'基本信息',{'fields':(('registerId','name','idstr'),('gender','age','phone',))}),
        ("1", {"classes": ("placeholder cha2ds2vacsevaluation_set-group",), "fields" : ()}),
        ("2", {"classes": ("placeholder hasbledevaluation_set-group",), "fields" : ()}),
#        [u'术前检查-UCG', {'fields':('LV','LA','RV','RA','EF')}],
#        [u'术前检查-TEE',{'fields':('hasThrombus','laa0Diameter')}],
        ('',{'fields':('af','otherDisease','lifeQualityScore')}),
        (u'术前检查',{"classes": ("placeholder precheck_set-group",), "fields" : ()}),
        ('',{'fields':(('opDate','opType','opMethod',),)}),
        (u'术中信息',{"classes": ("placeholder insurgery_set-group",), "fields" : ()}),
        (u'术后用药',{'fields': ('anticoagulant', 'antiarrhythmic',)}),
        (u'术后随访',{"classes": ("placeholder surgeryfollow_set-group",), "fields" : ()}),

    )
    filter_horizontal = (
    #    'anticoagulant', 'antiarrhythmic',
    )
    list_display = (
        'registerId', 'name','idstr','gender','age','phone','af',
    )

    verbose_name = u'患者信息'
    verbose_name_plural = u'患者信息'
    actions = [export_xlsx]


# Register your models here.
class MyAdminSite(admin.AdminSite):
    site_header = u'信息统计'  # 此处设置页面显示标题
    site_title = u'手术信息统计系统'  # 此处设置页面头部标题
 
admin_site = MyAdminSite(name='stat')
#admin.site.register(PatientBasic,PatientAdmin)
admin_site.register(PatientBasic,PatientAdmin)
admin_site.register(SurgeryMethod)
admin_site.register(AblationType)
admin_site.register(Anticoagulant)
admin_site.register(Antiarrhythmic)
#admin.site.register([SurgeryType])
